# -*- coding: utf-8 -*-
"""
Created on Wed Sep  1 21:22:23 2021

@author: Alexander Mikhailov
"""

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from .core.config import DATA_DIR, DATA_EXTERNAL_PATH
from .utils.filenames import url_to_archive_name


def get_archive_names(file_name: str, path_src: Path = DATA_DIR) -> set[str]:
    """
    Extract archive names from the 'ref' column of an Excel file,
    converting URLs to standardized archive filenames.

    Parameters
    ----------
    file_name : str
        The name of the Excel file to read.
    path_src : Path, optional
        Directory containing the Excel file. Defaults to DATA_DIR.

    Returns
    -------
    set[str]
        A set of archive filenames.
    """
    archive_names: set[str] = set()
    workbook = load_workbook(path_src / file_name, read_only=True)
    sheet = workbook.active

    headers = [cell for cell in next(sheet.iter_rows(values_only=True))]
    try:
        ref_idx = headers.index('ref')
    except ValueError as exc:
        raise ValueError("No 'ref' column found in the Excel file") from exc

    for row in sheet.iter_rows(min_row=2, values_only=True):
        url = row[ref_idx]
        if url is None:
            continue
        try:
            archive_name = url_to_archive_name(url)
            archive_names.add(archive_name)
        except (IndexError, ValueError):
            continue

    return archive_names


def snapshot_sort_key(filepath: Path):
    """
    Extracts date from filename for sorting.
    Expected format: snapshot_YYYY-MM-DD.xlsx
    """
    try:
        date_str = filepath.stem.split('_')[-1]
        return datetime.strptime(date_str, '%Y-%m-%d')
    except Exception:
        # Fallback: lexical sort if parsing fails
        return filepath.name


def get_latest_snapshot(files: list[Path]):
    return files[-1] if files else None


def get_previous_snapshot(files: list[Path]):
    return files[-2] if len(files) >= 2 else None


def main(
    check_option: str,
    path_src: Path = DATA_DIR,
    url_root: str = 'https://www150.statcan.gc.ca/n1/tbl/csv',
):
    """
    Parameters
    ----------
    check_option : str
        'snapshots': Check Two Excel Files;
        'downloaded':
            Check the Latest Excel File Against Downloaded Collection.
    path_src : Path, optional
        Directory containing snapshot Excel files. Default is DATA_DIR.
    url_root : str, optional
        Base URL for StatCan archives.

    Returns
    -------
    None
    """

    snapshots_available = [
        f for f in path_src.iterdir() if f.suffix == '.xlsx'
    ]
    snapshots_available.sort(key=snapshot_sort_key)

    if not snapshots_available:
        print(f'No snapshot files found in {path_src}')
        return

    latest_snapshot = get_latest_snapshot(snapshots_available)
    previous_snapshot = get_previous_snapshot(
        snapshots_available) if check_option == 'snapshots' else None

    if check_option == 'snapshots' and not previous_snapshot:
        print('Not enough snapshots to compare (need at least 2).')
        return

    archive_names_available = get_archive_names(latest_snapshot)

    if check_option == 'snapshots':
        archive_names_seen = get_archive_names(previous_snapshot)
    else:
        archive_names_seen = {
            f.name for f in DATA_EXTERNAL_PATH.iterdir()
            if f.suffix == '.zip' and f.name.endswith('-eng.zip')
        }

    archive_names_to_check = sorted(
        archive_names_available - archive_names_seen)

    if archive_names_to_check:
        print('You Might Want to Check Those New Archives:')
        for archive_name in archive_names_to_check:
            print(f'{url_root}/{archive_name}')
    else:
        print('No New Archives Since the Last Snapshot/Download')


if __name__ == '__main__':
    main('snapshots')
