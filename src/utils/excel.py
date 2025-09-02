from pathlib import Path

from openpyxl import load_workbook

from .filenames import url_to_archive_name


def extract_archive_names_from_excel(file_path: Path) -> set[str]:
    """
    Extract archive names from the 'ref' column of an Excel file,
    converting URLs to standardized archive filenames.
    """
    archive_names: set[str] = set()
    workbook = load_workbook(file_path, read_only=True)
    sheet = workbook.active

    headers = list(next(sheet.iter_rows(values_only=True)))
    try:
        ref_column_index = headers.index('ref')
    except ValueError as exc:
        raise ValueError("No 'ref' column found in the Excel file") from exc

    for row in sheet.iter_rows(min_row=2, values_only=True):
        url = row[ref_column_index]
        if url is None:
            continue
        try:
            archive_name = url_to_archive_name(url)
            archive_names.add(archive_name)
        except (IndexError, ValueError):
            continue

    return archive_names
